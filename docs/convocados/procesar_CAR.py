"""
Procesa CAR.xlsx y genera:
  CAR_nuevos.sql       — INSERT de jugadores nuevos
  CAR_posiciones.sql   — UPDATE de posición para titulares con FORWARD/BACK
  CAR_convocados.csv   — id,plantel para importar
  CAR_reporte.txt      — reporte de matches, ambiguos, nuevos, suplentes-only
"""

import openpyxl
import psycopg2
import unicodedata
import os

DB_URL = "postgresql://postgres:postgres@localhost:5432/grantrico"
XLSX = os.path.join(os.path.dirname(__file__), "CAR.xlsx")
OUT = os.path.dirname(__file__)

POSITION_MAP = {
    1: 'PILAR', 2: 'HOOKER', 3: 'PILAR',
    4: 'SEGUNDA_LINEA', 5: 'SEGUNDA_LINEA',
    6: 'TERCERA_LINEA', 7: 'TERCERA_LINEA', 8: 'TERCERA_LINEA',
    9: 'MEDIO_SCRUM', 10: 'APERTURA',
    11: 'WING', 12: 'CENTRO', 13: 'CENTRO', 14: 'WING', 15: 'FULLBACK',
}

PLANTELES = {1: 'PRIMERA', 3: 'INTER', 5: 'PRE_A', 7: 'PRE_B', 9: 'PRE_C', 11: 'PRE_D'}

# Matches manuales: (apellido_norm_de_xlsx, nombre_norm_parcial) -> jugador_id
# Cubre apellidos compuestos, tildes distintas, y cualquier ambigüedad
MANUAL_MATCHES = {
    ('gianoli', 'gianluca'): 151,       # Gian Gianoli
    ('gianoli', 'marco'): 188,          # Marco Gianoli
    ('layno', 'gonzalo'): 183,          # Gonzalo Layño
    ('laino', 'gonzalo'): 183,
    ('ghizzoli', 'juan'): 196,          # Juan Ignacio Ghizzoni
    ('ghizzoni', 'juan'): 196,
    ('siragusa', 'gonzalo'): 218,       # Gonzalo Sigarusa
    ('sigarusa', 'gonzalo'): 218,
    ('forastiero', 'ivan'): 224,        # Ivan Panzitta
    ('panzitta', 'ivan'): 224,
    ('berdeski', 'ramiro'): 256,        # Ramiro Berdesky
    ('berdesky', 'ramiro'): 256,
    ('di pascuale', 'juan'): 239,       # Juan Manuel Di Pascuale
    ('dibacco', 'felipe'): 215,
    ('dibacco', ''): 215,               # "Dibacco" solo
    # Apellidos compuestos
    ('bargiela', 'santiago'): 178,      # Santiago Rodriguez Bargiela
    ('ricci', 'ignacio'): 165,          # Ignacio Rodriguez Ricci
    ('badia', 'lautaro'): 192,          # Lautaro Lopez Badia
    ('cafiero', 'velentin'): 249,       # Velentin campos Cafiero
}

# Variantes adicionales de apellido que pueden venir del xlsx
APELLIDO_ALIASES = {
    'ghizzoli': 'ghizzoni',
    'laiño': 'layño',
    'laino': 'layno',
    'siragusa': 'sigarusa',
    'berdeski': 'berdesky',
}


def normalize(s):
    """Elimina tildes, lowercase, strip."""
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.lower().strip()


def parse_name(raw):
    """Devuelve (apellido, nombre) del string. Varios formatos posibles."""
    raw = raw.replace('( C)', '').replace('(C)', '').strip()
    return raw  # devuelve el string completo; el split se hace al buscar


def load_players_from_db():
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    cur.execute('SELECT id, nombre, apellido, apodo, posicion FROM "Jugador"')
    rows = cur.fetchall()
    conn.close()

    players = []
    by_apellido = {}  # apellido_norm -> list of player dicts
    for id_, nombre, apellido, apodo, posicion in rows:
        p = {'id': id_, 'nombre': nombre, 'apellido': apellido, 'apodo': apodo, 'posicion': posicion}
        players.append(p)
        key = normalize(apellido)
        by_apellido.setdefault(key, []).append(p)
    return players, by_apellido


def find_player(raw_name, db_by_apellido):
    """
    Intenta encontrar el jugador en la DB.
    Devuelve (player_dict, method) o (None, None).
    """
    raw = raw_name.replace('( C)', '').replace('(C)', '').strip()
    parts = raw.split()

    # Intentar como "Nombre Apellido" y "Apellido Nombre"
    if not parts:
        return None, None

    # Intento 1: última palabra como apellido
    apellido_raw = parts[-1]
    nombre_raw = ' '.join(parts[:-1]) if len(parts) > 1 else ''

    apellido_norm = normalize(apellido_raw)
    apellido_norm_alias = APELLIDO_ALIASES.get(apellido_norm, apellido_norm)
    nombre_norm = normalize(nombre_raw)

    # Check manual matches primero
    for (ap, nom), pid in MANUAL_MATCHES.items():
        if (apellido_norm == ap or apellido_norm_alias == ap) and (not nom or nombre_norm.startswith(nom) or nom.startswith(nombre_norm)):
            # Buscar el player con este id
            for key, players in db_by_apellido.items():
                for p in players:
                    if p['id'] == pid:
                        return p, 'manual'

    # Intento: solo apellido como key (última palabra)
    candidates = db_by_apellido.get(apellido_norm_alias, db_by_apellido.get(apellido_norm, []))
    if len(candidates) == 1:
        return candidates[0], 'apellido'
    elif len(candidates) > 1 and nombre_norm:
        for c in candidates:
            if normalize(c['nombre']).startswith(nombre_norm[:3]):
                return c, 'apellido+nombre'

    # Intento 2: apellido compuesto (primeras dos palabras)
    if len(parts) >= 2:
        apellido_comp = normalize(parts[0] + ' ' + parts[1])
        candidates2 = db_by_apellido.get(apellido_comp, [])
        if len(candidates2) == 1:
            return candidates2[0], 'apellido_compuesto'

    return None, None


def parse_xlsx():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    titulares = []   # (jersey, name, plantel, posicion)
    suplentes_raw = []  # (name, plantel)

    seen_suplentes = set()

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        for col_idx, plantel in PLANTELES.items():
            num_val = row[col_idx - 1] if col_idx - 1 < len(row) else None
            name_val = row[col_idx] if col_idx < len(row) else None
            if not name_val:
                continue
            name = str(name_val).strip()
            if not name or name == 'None':
                continue
            try:
                jersey = int(str(num_val).strip()) if num_val else 0
            except (ValueError, AttributeError):
                continue

            if jersey <= 15:
                titulares.append((jersey, name, plantel, POSITION_MAP.get(jersey, 'FORWARD')))
            else:
                key = name.lower().strip()
                if key not in seen_suplentes:
                    seen_suplentes.add(key)
                    suplentes_raw.append((name, plantel))

    return titulares, suplentes_raw


def main():
    print("Cargando jugadores de la DB...")
    all_players, by_apellido = load_players_from_db()
    print(f"  {len(all_players)} jugadores en DB")

    print("Parseando xlsx...")
    titulares, suplentes_raw = parse_xlsx()

    titular_names_norm = {normalize(t[1]) for t in titulares}

    # Suplente-only: no aparecen como titular en ningún plantel
    suplentes_only = [(name, plantel) for name, plantel in suplentes_raw
                      if normalize(name) not in titular_names_norm]

    print(f"  {len(titulares)} titulares, {len(suplentes_only)} suplentes-only")

    # === Matching ===
    matched = []       # (name, plantel, posicion, player_id, method)
    new_players = []   # (name, plantel, posicion, jersey)
    ambiguos = []

    # Procesar titulares
    for jersey, name, plantel, posicion in titulares:
        player, method = find_player(name, by_apellido)
        if player:
            matched.append((name, plantel, posicion, player['id'], player['posicion'], method))
        else:
            new_players.append((name, plantel, posicion, jersey, 'titular'))

    # Procesar suplentes-only
    for name, plantel in suplentes_only:
        player, method = find_player(name, by_apellido)
        if player:
            matched.append((name, plantel, 'FORWARD', player['id'], player['posicion'], method + '_sup'))
        else:
            new_players.append((name, plantel, 'FORWARD', 0, 'suplente_only'))

    # === Generar SQL de nuevos jugadores ===
    sql_nuevos = []
    sql_nuevos.append("-- Nuevos jugadores para CAR fecha 6")
    sql_nuevos.append("INSERT INTO \"Jugador\" (nombre, apellido, apodo, posicion, activo) VALUES")

    new_entries = []
    for name, plantel, posicion, jersey, tipo in new_players:
        parts = name.split()
        # Heurística: si el nombre tiene 2 palabras → nombre apellido
        # si tiene 3+ → primeras palabras = nombre, última = apellido
        if len(parts) == 1:
            nombre_db = ''
            apellido_db = parts[0]
        elif len(parts) == 2:
            nombre_db = parts[0]
            apellido_db = parts[1]
        else:
            # "Juan Ignacio Ghizzoli" → nombre="Juan Ignacio", apellido="Ghizzoli"
            apellido_db = parts[-1]
            nombre_db = ' '.join(parts[:-1])

        apodo = apellido_db.upper()
        new_entries.append(f"  ('{nombre_db}', '{apellido_db}', '{apodo}', '{posicion}', true)  -- {name} [{plantel}] tipo={tipo}")

    sql_nuevos.append(',\n'.join(new_entries) + ';')

    with open(os.path.join(OUT, 'CAR_nuevos.sql'), 'w') as f:
        f.write('\n'.join(sql_nuevos))
    print(f"  -> CAR_nuevos.sql ({len(new_players)} jugadores nuevos)")

    # === Generar SQL de actualización de posiciones ===
    sql_pos = []
    sql_pos.append("-- Actualización de posiciones titulares CAR fecha 6")
    updates = 0
    for name, plantel, posicion, pid, old_posicion, method in matched:
        if old_posicion in ('FORWARD', 'BACK') and posicion not in ('FORWARD', 'BACK') and '_sup' not in method:
            sql_pos.append(f"UPDATE \"Jugador\" SET posicion = '{posicion}' WHERE id = {pid};  -- {name} ({old_posicion} -> {posicion})")
            updates += 1

    with open(os.path.join(OUT, 'CAR_posiciones.sql'), 'w') as f:
        f.write('\n'.join(sql_pos) if sql_pos else '-- No hay actualizaciones necesarias\n')
    print(f"  -> CAR_posiciones.sql ({updates} actualizaciones de posición)")

    # === Reporte ===
    report = []
    report.append("=== REPORTE DE MATCHING CAR ===\n")

    report.append(f"--- MATCHES ({len(matched)}) ---")
    for name, plantel, posicion, pid, old_pos, method in matched:
        flag = " [ACTUALIZA POSICION]" if old_pos in ('FORWARD', 'BACK') and posicion not in ('FORWARD', 'BACK') and '_sup' not in method else ""
        report.append(f"  [{plantel}] {name} -> id={pid} ({method}){flag}")

    report.append(f"\n--- JUGADORES NUEVOS ({len(new_players)}) ---")
    for name, plantel, posicion, jersey, tipo in new_players:
        report.append(f"  [{plantel}] {name} -> {posicion} (tipo={tipo})")

    report.append(f"\n--- SUPLENTES-ONLY ({len(suplentes_only)}) ---")
    for name, plantel in suplentes_only:
        player, method = find_player(name, by_apellido)
        status = f"match id={player['id']} ({method})" if player else "NUEVO"
        report.append(f"  [{plantel}] {name} -> {status}")

    with open(os.path.join(OUT, 'CAR_reporte.txt'), 'w') as f:
        f.write('\n'.join(report))
    print("  -> CAR_reporte.txt")

    # === CSV de convocados (para jugadores ya en DB) ===
    # Nota: los nuevos jugadores no tienen id todavía, el CSV se genera después de aplicar el SQL
    csv_lines = ["id,plantel"]
    for name, plantel, posicion, pid, old_pos, method in matched:
        csv_lines.append(f"{pid},{plantel}")

    with open(os.path.join(OUT, 'CAR_convocados_existentes.csv'), 'w') as f:
        f.write('\n'.join(csv_lines))
    print(f"  -> CAR_convocados_existentes.csv ({len(matched)} jugadores)")

    print("\nDone. Ahora:")
    print("  1. Revisar CAR_reporte.txt")
    print("  2. Revisar CAR_nuevos.sql y ajustar nombres/apodos si hace falta")
    print("  3. psql ... < CAR_nuevos.sql")
    print("  4. psql ... < CAR_posiciones.sql")
    print("  5. Correr procesar_CAR_step2.py para generar CSV final con los nuevos ids")


if __name__ == '__main__':
    main()
